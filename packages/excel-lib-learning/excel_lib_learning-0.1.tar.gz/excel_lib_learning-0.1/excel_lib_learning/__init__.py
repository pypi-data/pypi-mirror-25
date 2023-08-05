from openpyxl import load_workbook
from openpyxl import Workbook
import os.path

main_file_name = None

wb = None

ws = None

def open_excel_file(name = 'profit.xlsx'):
	global main_file_name
	global wb
	global ws
	main_file_name = name
	if not os.path.isfile(main_file_name):
		wb = Workbook()
	else:
		wb = load_workbook(main_file_name)
	ws = wb.active

def add_to_excel(cell, value):
	global ws
	ws[cell] = value

def get_from_excel(cell):
	global ws
	return ws[cell].value

def save_excel():
	global wb
	wb.save(main_file_name)