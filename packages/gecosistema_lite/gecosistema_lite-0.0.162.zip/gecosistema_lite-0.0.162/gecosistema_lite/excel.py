# -------------------------------------------------------------------------------
# Licence:
# Copyright (c) 2012-2017 Luzzi Valerio 
#
# The above copyright notice and this permission notice shall be
# included in all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND,
# EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES
# OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND
# NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT
# HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY,
# WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING
# FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR
# OTHER DEALINGS IN THE SOFTWARE.
#
#
# Name:        excel.py
# Purpose:
#
# Author:      Luzzi Valerio
#
# Created:     14/08/2017
# -------------------------------------------------------------------------------
import xlrd, xlwt
from xlutils.copy import copy
import openpyxl
from filesystem import *

XLRD = 0
OPENPYXL = 1

DATATYPES = {
    xlrd.XL_CELL_EMPTY: openpyxl.cell.Cell.TYPE_NULL,
    xlrd.XL_CELL_TEXT: openpyxl.cell.Cell.TYPE_STRING,
    xlrd.XL_CELL_NUMBER: openpyxl.cell.Cell.TYPE_NUMERIC,
    xlrd.XL_CELL_DATE: openpyxl.cell.Cell.TYPE_STRING,
    xlrd.XL_CELL_BOOLEAN: openpyxl.cell.Cell.TYPE_BOOL,
    xlrd.XL_CELL_ERROR: openpyxl.cell.Cell.TYPE_ERROR,
    xlrd.XL_CELL_BLANK: openpyxl.cell.Cell.TYPE_NULL,

    openpyxl.cell.Cell.TYPE_NULL: xlrd.XL_CELL_EMPTY,
    openpyxl.cell.Cell.TYPE_STRING: xlrd.XL_CELL_TEXT,
    openpyxl.cell.Cell.TYPE_NUMERIC: xlrd.XL_CELL_NUMBER,
    openpyxl.cell.Cell.TYPE_BOOL: xlrd.XL_CELL_BOOLEAN,
    openpyxl.cell.Cell.TYPE_ERROR: xlrd.XL_CELL_ERROR
}


class Workbook:

    def __init__(self, filename='', type='xls', mode='r'):
        """
        constructor
        """
        self.lib = OPENPYXL if type == "xlsx" else XLRD

        if not filename:
            if self.lib == OPENPYXL:
                self.wb = openpyxl.Workbook()
            elif self.lib == XLRD:
                self.wb = xlwt.Workbook()

        elif mode == "r":
            if self.lib == OPENPYXL:
                self.wb = openpyxl.load_workbook(filename, read_only=True)
            elif self.lib == XLRD:
                self.wb = xlrd.open_workbook(filename, on_demand=True)

        elif mode == "rw":
            if self.lib == OPENPYXL:
                self.wb = openpyxl.load_workbook(filename)
            elif self.lib == XLRD:
                rb = xlrd.open_workbook(filename, on_demand=True)
                self.wb = copy(rb)

    def open_workbook(self, filename):
        pass

    def add_sheet(self, sheet_name):
        """
        add_sheet
        :param sheet_name:
        :return:
        """
        if self.lib == OPENPYXL:
            sheet = self.wb.create_sheet(title=sheet_name)
        elif self.lib == XLRD:
            sheet = self.wb.add_sheet(sheet_name)
        return WorkSheet(sheet)

    def create_sheet(self, sheet_name):
        return self.add_sheet(sheet_name)

    def sheets(self):
        """
        sheets
        :return:
        """
        if self.lib == OPENPYXL:
            return [WorkSheet(sheet) for sheet in self.wb]
        elif self.lib == XLRD:
            return [WorkSheet(sheet) for sheet in self.wb.sheets()]
        return []

    def sheet_names(self):
        """
        sheet_names
        :return:
        """
        if self.lib == OPENPYXL:
            return self.wb.get_sheet_names()
        elif self.lib == XLRD:
            return self.wb.sheet_names()
        return []

    def get_sheets_by_names(self):
        return self.sheet_names()

    def sheet_by_index(self, index):
        """
        sheet_by_index
        :param index:
        :return:
        """
        if self.lib == OPENPYXL:
            ws = self.wb.sheetnames[index]
        elif self.lib == XLRD:
            ws = self.wb.sheet_by_index(index)
        else:
            ws = None
        return WorkSheet(ws) if ws else None

    def get_sheet(self, index):
        return self.sheet_by_index(index)

    def sheet_by_name(self, sheet_name):
        """
        get_sheet_by_name
        :param sheet_name:
        :return:
        """
        if self.lib == OPENPYXL:
            ws = self.wb.get_sheet_by_name(sheet_name)
        elif self.lib == XLRD:
            ws = self.wb.sheet_by_name(sheet_name)
        else:
            ws = None
        return WorkSheet(ws) if ws else None

    def get_sheet_by_name(self, sheet_name):
        return self.sheet_by_name(sheet_name)

    def save(self, filename):
        """
        save
        :param filename:
        :return:
        """
        if self.lib == OPENPYXL:
            self.wb.save(filename)
        elif self.lib == XLRD:
            self.wb.save(filename)


class WorkSheet:
    def __init__(self, sheet):
        if isinstance(sheet, (xlrd.sheet.Sheet, xlwt.Worksheet)):
            self.lib = XLRD
            # self.nrows = sheet.nrows
            # self.ncols = sheet.ncols
        else:
            self.lib = OPENPYXL
            # self.nrows = sheet.max_row
            # self.ncols = sheet.max_column
        self.ws = sheet

    def cell(self, i, j, value=None):
        if self.lib == OPENPYXL:
            if value == None:
                cl = self.ws.cell(row=i + 1, column=j + 1)
            else:
                # write mode
                cl = self.ws.cell(row=i + 1, column=j + 1, value=value)
        elif self.lib == XLRD:
            if value == None:
                cl = self.ws.cell(rowx=i, colx=j)
            else:
                cl = self.ws.write(i, j, value)
                # cl = self.ws.cell(rowx=i, colx=j)
        else:
            cl = None

        return cl

    def cell_value(self, i, j):
        cl = self.cell(i, j)
        return cl.value if cl else None

    def cell_type(self, i, j):
        cl = self.cell(i, j)
        return cl.ctype if cl else None

    def write(self, i, j, value):
        if self.lib == OPENPYXL:
            self.ws.cell(i, j, value)
        elif self.lib == XLRD:
            self.ws.write(i, j, value)

    def merged_cells(self):
        if self.lib == OPENPYXL:
            return self.ws.merged_cell_ranges
        elif self.lib == XLRD:
            return self.ws.merged_cells
        return []

    def get_rows(self):
        if self.lib == OPENPYXL:
            return self.ws.rows
        elif self.lib == XLRD:
            return self.ws.get_rows()
        return []

    def get_columns(self):
        if self.lib == OPENPYXL:
            return self.ws.columns
        elif self.lib == XLRD:
            return self.ws.get_columns()
        return []


if __name__ == "__main__":
    filename = "./tests/xls/Dati_stazione_superficiale_test.xlsx"
    from gecosistema_lite import *

    #db = SqliteDB.From(filename, "Foglio1")
    from gecosistema_lite import *
    import xlwt

    filename1 = "./tests/xls/test2.xls"  # using xlrd

    #wb = xlwt.Workbook()
    wb = Workbook()
    print wb

    sheet = wb.add_sheet("Foglio4")

    sheet = wb.add_sheet("Foglio5")
    print wb.sheet_by_index(0)
