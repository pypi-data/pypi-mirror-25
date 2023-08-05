#Version 2.5
##Does not use win32com

#Import standard elements
import os
import sys
import subprocess
from ctypes import windll as ctypesWindll #Used to determine screen dpi

#Import the openpyxl module to work with excel sheets
import openpyxl

#Import needed support modules
import PIL.ImageGrab

class Utilities():
	def __init__(self):
		"""Functions to make the Excel module easier.

		Example Input: Utilities()
		"""

		#Internal Variables
		self.utilitiesDebugging = False

	def convertColumn(self, column):
		"""Converts a column number to a column letter, and returns it to the user as a string.

		column (int)  - The index of the column

		Example Input: convertColumn(3)
		"""

		#Convert Column if needed
		if (type(column) == int):
			#Check for past Z
			count = 0
			bonusColumn = ""
			while True:
				count += 1
				#Does the askii letter go past Z? If so, create addition letter
				if (openpyxl.utils.get_column_letter(count).isupper()):
					break
				else:
					column -= 26
					bonusColumn = openpyxl.utils.get_column_letter(count)

			#Set new Column
			column = bonusColumn + openpyxl.utils.get_column_letter(column)

			return column
		return None

class Excel(Utilities):
	def __init__(self):
		"""Works with excel files.
		Documentation for openpyxl can be found at: https://openpyxl.readthedocs.io/en/default/index.html

		Example Input: Excel()
		"""
		super(Excel, self).__init__()
		
		self.newBook("newBook")
		
		self.imageCatalogue = {} #(dict) - Used to catalogue all of the images in the document. {sheet title: [top-left corner cell (row, column), image as a PIL image]}
		self.excelDebugging = True

		# for i, row in enumerate(self.sheet.iter_cols()):

		# for i, column in enumerate(self.sheet.iter_rows()):

	def newBook(self, title):
		"""Creates a new workbook ans saves it in memmory.

		title (str) - The title of the workbook

		Example Input: newBook("test")
		"""

		self.book = openpyxl.Workbook()
		self.setBookTitle(title)
		self.sheet = self.book.active

	def setBookTitle(self, title):
		"""Changes the title of the workbook.

		title (str) - The title of the workbook

		Example Input: setBookTitle("test")
		"""

		self.book.title = title

	def newSheet(self, position = None, title = None, tabColor = None, changeToSheet = True):
		"""Adds a new sheet to the excel file.

		position (int)       - Where to insert the sheet at. If None: Insert at the end
		title (str)          - The name of the sheet. If None: It is given the default name (ie: Sheet, Sheet1, Sheet2, etc.)
		tabColor (str)       - The RRGGBB color code for the tab. If None: it is the default white 
		changeToSheet (bool) - Wether to change the current sheet to this new sheet or not

		Example Input: newSheet()
		Example Input: newSheet(0, "Aimer", "1072BA")
		"""

		if (changeToSheet):
			if (position != None):
				self.sheet = self.book.create_sheet(position)
			else:
				self.sheet = self.book.create_sheet()
		else:
			if (position != None):
				self.book.create_sheet(position)
			else:
				self.book.create_sheet()


		if (title != None):
			self.setSheetTitle(title)

		if (tabColor != None):
			self.setSheetTabColor(tabColor)

	def deleteSheet(self, sheetName):
		"""Removes a sheet from the book.

		sheetName (str) - The name of the sheet to be removed from the book

		Example Input: deleteSheet("sheet1")
		"""

		#Get the sheet
		sheet = self.getSheet(sheetName)

		#Remove the sheet
		self.book.remove_sheet(sheet)

	def setSheetTitle(self, name):
		"""Changes the name of a sheet.

		title (str) - The name of the sheet

		Example Input: setSheetTitle("test")
		"""

		self.sheet.title = name

	def getSheetTitle(self):
		"""Returns the name of the current sheet.

		Example Input: getSheetTitle()
		"""

		title = self.sheet.title
		return title

	def setSheetTabColor(self, tabColor):
		"""Changes color of a sheet's tab.

		tabColor (str) - The RRGGBB color code for the tab. If None: it is the default white 

		Example Input: setSheetTabColor("test")
		"""

		self.sheet.sheet_properties.tabColor = tabColor

	def getSheet(self, which):
		"""Returns a sheet when given the sheet's name or index number.
		Returns the desired sheet.

		which (str) - The name of the desired sheet. Can be the index as an int

		Example Input: getSheet("Aimer")
		"""

		if (type(which) == str):
			sheet = self.book.get_sheet_by_name(which)

		else:
			sheet = self.book.worksheets[which]

		return sheet

	def getAllSheetNames(self):
		"""Returns a list of all the sheet names as strings."""

		return self.book.get_sheet_names() #Example Return: ['Sheet2', 'New Title', 'Sheet1']

	def changeSheet(self, which):
		"""Changes to a sheet when given the sheet's name, index number, or the sheet itelf.

		which (str) - The name of the desired sheet. Can be an index ans an int or a sheet object.

		Example Input: changeSheet("Aimer")
		Example Input: changeSheet(0)
		Example Input: changeSheet(sheet)
		"""

		if ((type(which) == str) or (type(which) == int)):
			self.sheet = self.getSheet(which)
		else:
			self.sheet = which

	#Getters
	def getCell(self, row, column):
		"""Returns a specific cell object.
		The top-left corner is row (1, 1) not (0, 0).

		row (int)    - The index of the row
		column (int) - The index of the column. Can be a char

		Example Input: getCell(1, 2)
		"""

		if (type(column) == int):
			column = openpyxl.utils.get_column_letter(column)

		return self.sheet[column + str(row)]

	def getCellValue(self, row = None, column = None, cell = None):
		"""Returns the contents of a cell.
		The top-left corner is row (1, 1) not (0, 0).

		cell (object) - An openpyxl cell object
		row (int)    - The index of the row
		column (int) - The index of the column. Can be a char

		Example Input: getCellValue(1, 2)
		Example Input: getCellValue(cell = myCell)
		"""

		if (cell != None):
			if ((row == None) and (column == None)):
				return cell.value
		else:
			if ((row != None) and (column != None)):
				if (type(column) == int):
					column = openpyxl.utils.get_column_letter(column)

				return self.sheet[column + str(row)].value

		print("Input Values are configured incorrectly.")
		print("Define only (A) row and column, or (B) cell.")
		return None

	def getCellImage(self, row, column):
		"""Returns a PIL image object from a cell. 
		Returns 'None' if no image was found on the sheet.
		The top-left corner is row (1, 1) not (0, 0).

		row (int)    - The index of the row
		column (int) - The index of the column

		Example Input: getCellImage(1, 2)
		"""

		#Get the sheet title
		title = self.getSheetTitle()

		#Get the catalogued images [coordinates (column is a letter), image]
		if (title in self.imageCatalogue):
			imageList = self.imageCatalogue[title]
		else:
			return None

		#Setup the search location coordinates as a tuple (like the imageCatalogue has)
		coordinates = (row, column)

		#Find the desired image
		image = None
		for item in imageList: 
			if (coordinates == item[0]):
				image = item[1]
				break

		return image

	#Setters
	def setCell(self, value, row, column):
		"""Writes the value of a cell.
		The top-left corner is row (1, 1) not (0, 0).

		value (any)    - What will be written to the cell
		row (int)    - The index of the row
		column (int) - The index of the column. Can be a char

		Example Input: setCell(42, 1, 2)
		Example Input: setCell(3.14, 1, "B")
		Example Input: setCell("Hello World", 1, 2)
		"""

		#Convert Column if needed
		if (type(column) == int):
			#Check for past Z
			count = 0
			bonusColumn = ""
			while True:
				count += 1
				#Does the askii letter go past Z? If so, create addition letter
				if (openpyxl.utils.get_column_letter(count).isupper()):
					break
				else:
					column -= 26
					bonusColumn = openpyxl.utils.get_column_letter(count)

			#Set new Column
			column = bonusColumn + openpyxl.utils.get_column_letter(column)

		#Write Value
		self.sheet[column + str(row)] = value

	def appendRow(self, contents = None):
		"""Appends a row to the end of the file.

		contents (list) - What the cells in the row will contain. If None, the row will be blank

		Example Input: appendRow()
		Example Input: appendRow([0, 1, 2, 3, 4, 5])
		"""

		#Find the last row
		row = len(list(self.sheet.iter_rows())) + 1

		#Write to cells
		if ((contents != None) and (len(contents) != 0)):
			for column, item in enumerate(contents):
				self.setCell(item, row, column + 1)
		else:
			self.setCell(" ", row, 1)

	def appendColumn(self, contents = None):
		"""Appends a column to the end of the file.

		contents (list) - What the cells in the column will contain. If None, the column will be blank

		Example Input: appendColumn()
		Example Input: appendColumn([0, 1, 2, 3, 4, 5])
		"""

		#Find the last column
		column = len(list(self.sheet.iter_cols())) + 1

		#Write to cells
		if ((contents != None) and (len(contents) != 0)):
			for row, item in enumerate(contents):
				self.setCell(item, row + 1, column)
		else:
			self.setCell(" ", column, 1)

	def setCellFormula(self, row, column, formula, *args):
		"""Writes an excel formula value to a cell.
		The top-left corner is row (1, 1) not (0, 0).

		row (int)    - The index of the row
		column (int) - The index of the column. Can be a char
		formula (str)  - The excel formula
		args*          - The arguments for the formula

		Example Input: setCell(1, 2, "SUM", 1, 2, 3, 4, 5, 6, 7)
		"""

		#Ensure formula format
		formula = formula.upper()

		#Check if formula exists
		if (formula in openpyxl.utils.FORMULAE):
			#Format args
			formattedArgs = ""
			for i, item in enumerate(args):
				formattedArgs += str(item) + ", "

			#Convert Column if needed
			if (type(column) == int):
				column = openpyxl.utils.get_column_letter(column)

			#Write to cell
			self.sheet[column + str(row)] = "=" + formula + formattedArgs

		else:
			print("ERROR: formula does not exist")

	def setCellImage(self, imagePath, row, column, xSize = None, ySize = None, keepAspectRatio = True):
		"""Inserts an image to a cell.
		The top-left corner is row (1, 1) not (0, 0).

		imagePath (str) - The path to the image. Can be a PIL image
		row (int)       - The index of the row
		column (int)    - The index of the column. Can be a char
		xSize (int)     - The width (in pixels) of the image on the excel file. If None: Do not size to this
		ySize (int)     - The height (in pixels) of the image on the excel file. If None: Do not Size to this
		keepAspectRatio (bool) - If True: The image aspect ratio will be preserved. The re-size will go off of the largest side

		Example Input: setCellImage("test.jpg", 1, 2)
		Example Input: setCellImage(image, 1, 2)
		"""

		#Convert Column if needed
		if (type(column) != str):
			column = self.convertColumn(column)

		#Add Image
		image = openpyxl.drawing.image.Image(imagePath, size = (xSize, ySize), nochangeaspect = keepAspectRatio)
		self.sheet.add_image(image, column + str(row))

	def setCellStyle(self, row, column, font = None, bold = None, size = None):
		"""Changes the style of the text in a cell.
		The top-left corner is row (1, 1) not (0, 0).

		row (int)    - The index of the row
		column (int) - The index of the column. Can be a char
		font (str)   - The name of the font style. If None: the current font style will be used
		bold (bool)  - If True: bolds the text. If False: unbolds the text. If None: does not change the boldness of the text
		size (int)   - The size of the text in the cell. The size '11' is typical. Can be a string

		Example Input: setCellStyle(2, 3)
		Example Input: setCellStyle(2, 3, "calibri")
		Example Input: setCellStyle(2, 3, bold = True)
		"""

		#Get the cell whose text style will be modified
		cell = self.getCell(row, column)

		#Determine the font style
		if (font != None):
			#Ensure lower-case
			font = font.tolower()

			#Set the font style
			if (font == "default"):
				font = openpyxl.styles.Font(style = "Normal")

			elif (font == "calibri"):
				font = openpyxl.styles.Font(name = "Calibri",
					size = 11,
					bold = False,
					italic = False,
					vertAlign = None,
					underline = "none",
					strike = False,
					color = "FF000000")

			elif (font == "title"):
				font = openpyxl.styles.Font(style = "Title")

			elif (font == "headline1"):
				font = openpyxl.styles.Font(style = "Headline 1")

			elif (font == "headline2"):
				font = openpyxl.styles.Font(style = "Headline 2")

			elif (font == "headline3"):
				font = openpyxl.styles.Font(style = "Headline 3")

			elif (font == "headline4"):
				font = openpyxl.styles.Font(style = "Headline 4")

			elif (font == "calculation"):
				font = openpyxl.styles.Font(style = "Calculation")

			elif (font == "warningText"):
				font = openpyxl.styles.Font(style = "Warning Text")

			elif (font == "input"):
				font = openpyxl.styles.Font(style = "Input")

			elif (font == "output"):
				font = openpyxl.styles.Font(style = "Output")

			elif (font == "good"):
				font = openpyxl.styles.Font(style = "Good")

			elif (font == "bad"):
				font = openpyxl.styles.Font(style = "Bad")

			elif (font == "neutral"):
				font = openpyxl.styles.Font(style = "Neutral")

			else:
				print("Font style", font, "not found")
				font = openpyxl.styles.Font(style = "Normal")

		else:
			font = cell.font.copy()

		#Determine text boldness
		if (bold != None):
			if (bold):
				font = font.copy(bold = True)
			else:
				font = font.copy(bold = False)

		#Determine if the text should be italisized, not, or stay whatever it is
		if (bold != None):
			if (bold):
				font = font.copy(italic = True)
			else:
				font = font.copy(italic = False)

		#Determine if the text size should be changed
		if (size != None):
			font = font.copy(size = int(size))

		#Apply the style to the cell
		cell.font = font

	def mergeCells(self, startRow, startColumn, endRow, endColumn):
		"""Merges a range of cells.
		The top-left corner is row (1, 1) not (0, 0).

		startRow (int) - The index of the left end row of the cells to merge. Can be a char
		startColumn (int) - The index of the left end column of the cells to merge
		endRow (int) - The index of the right end row of the cells to merge. Can be a char
		endColumn (int) - The index of the left end column of the cells to merge

		Example Input: mergeCells(1, 2, 3, 2)
		"""

		if (type(startColumn) == str):
			self.sheet.mergeCells(startColumn + str(startRow) + ":" + endColumn + str(endRow))

		elif (type(endColumn) == int):
			self.sheet.merge_cells(start_row = startRow, start_column = startColumn, end_row = endRow, end_column = endColumn)

		else:
			print("ERROR: row type error")

	def unmergeCells(self, startRow, startColumn, endRow, endColumn):
		"""Unmerges a range of cells.
		The top-left corner is row (1, 1) not (0, 0).

		startRow (int) - The index of the left end row of the cells to unmerge. Can be a char
		startColumn (int) - The index of the left end column of the cells to unmerge
		endRow (int) - The index of the right end row of the cells to unmerge. Can be a char
		endColumn (int) - The index of the left end column of the cells to unmerge

		Example Input: unmergeCells(1, 2, 3, 2)
		"""

		if (type(startColumn) == str):
			self.sheet.unmergeCells(startColumn + str(startRow) + ":" + endColumn + str(endRow))

		elif (type(endColumn) == int):
			self.sheet.unmerge_cells(start_row = startRow, start_column = startColumn, end_row = endRow, end_column = endColumn)

		else:
			print("ERROR: row type error")

	def hideColumns(self, startColumn, endColumn):
		"""Hides a range of columns.

		startColumn (int) - The index of the left end column of the cells to hide. Can be a char
		endColumn (int)   - The index of the right end column of the cells to hide. Can be a char

		Example Input: hideColumns(1, 2)
		Example Input: hideColumns("A", "D")
		"""

		#Format Columns
		if (type(startColumn) == int):
			startColumn = openpyxl.utils.get_column_letter(startColumn)

		if (type(endColumn) == int):
			endColumn = openpyxl.utils.get_column_letter(endColumn)

		self.sheet.column_dimensions.group(startColumn, endColumn, hidden = True)

	def autosizeColumns(self):
		"""Autosizes all the columns of a sheet.

		Example Input: autosizeColumns()
		"""

		for i, row in enumerate(self.sheet.columns):
			if (i%3 == 0):
				self.setColumnColor(i + 3)
			self.setColumnWidth(i + 1)
		self.setColumnWidth(i + 2)

	def getColumn(self, column = None, justLength = False, contents = True, trailingNone = False, noNone = True):
		"""Returns a list of all openpyxl cell objects within the column
		Any blank cell has a value of None. All columns are the length of the longest column.
		To get just the desired column, make sure that trailingNone is True.

		column (int)       - Which column to retrieve (The first column is '1'). If None: Returns the maximum column
		justLength (bool)  - If True: Returns the number of occupied cells in the column
		contents (bool)    - If True: Returns the cell contents instead of the cell objects
		trailingNone(bool) - If False: Any blank cells will be removed from the end of the list
		noNone (bool)      - If True: Any internal blank cell will be returned as "" instead of None
							 Note: Only applies if 'contents' is True

		Example Input: getColumn()
		Example Input: getColumn(2)
		Example Input: getColumn(justLength = True)
		Example Input: getColumn(2, True)
		Example Input: getColumn(2, trailingNone = False)
		Example Input: getColumn(2, noNone = False)
		"""

		#Allow for string inputs
		if (type(column) == str):
			column = int(column)

		#Get all of the columns in the sheet
		allColumns = list(self.sheet.columns)

		#Determine if trailing None values should be stripped off
		modifiedColumns = []
		if (not trailingNone):
			for singleColumn in allColumns:
				singleColumn = list(singleColumn)
				while(len(singleColumn) > 0):
					if (self.getCellValue(cell = singleColumn[-1]) == None):
						singleColumn.pop(-1)
					else:
						break
				modifiedColumns.append(singleColumn)

			#Update the columns to not have None values)
			allColumns = modifiedColumns[:]

		#Determine if a specific column or the longest column is desired
		i = 0
		if (type(column) == int):
			#Fix it so that the column number is in the same format as the rest of the module
			i = column - 1
		else:
			#Get the length of each column, and then choose the largest one
			lengths = list(map(len, allColumns))
			i = np.argmax(lengths)

		item = allColumns[i]
		#print("@@@@@@", column, item)

		#Determine if the length or the list of cell objects should be returned
		if (justLength):
			return len(item)

		#Determine if the cell contents or cell objects should be returned
		if (contents):
			#Loop through each cell and retrieve its contents
			cellContents = []
			for piece in item:
				contents = self.getCellValue(cell = piece)
				cellContents.append(contents)

			#Determine if None values should be replaced with "".
			if (noNone):
				for i, piece in enumerate(cellContents):
					if (piece == None):
						cellContents[i] = ""

			return cellContents
		return item

	def getRow(self, row = None, justLength = False, contents = True, trailingNone = True, noNone = True):
		"""Returns a list of all openpyxl cell objects within the row

		row (int)     - Which row to retrieve (The first row is '1'). If None: Returns the maximum row
		justLength (int) - If True: Returns the number of occupied cells in the row
		contents (bool)   - If True: Returns the cell contents instead of the cell objects
		trailingNone(bool)  - If True: Any blank cells will be removed from the end of the list
		noNone (bool)      - If True: Any internal blank cell will be returned as "" instead of None
							 Note: Only applies if 'contents' is True

		Example Input: getRow()
		Example Input: getRow(2)
		Example Input: getRow(justLength = True)
		Example Input: getRow(2, True)
		Example Input: getRow(2, trailingNone = False)
		Example Input: getRow(2, noNone = False)
		"""

		#Allow for string inputs
		if (type(row) == str):
			row = int(row)

		#Get all of the rows in the sheet
		allRows = list(self.sheet.rows)

		#Determine if trailing None values should be stripped off
		modifiedRows = []
		if (not trailingNone):
			for singleRow in allColumns:
				singleRow = list(singleRow)
				while(len(singleRow) > 0):
					if (self.getCellValue(cell = singleRow[-1]) == None):
						singleRow.pop(-1)
					else:
						break
				modifiedRows.append(singleRow)

			#Update the rows to not have None values
			allColumns = modifiedRows[:]

		#Determine if a specific row or the longest row is desired
		i = 0
		if (type(row) == int):
			#Fix it so that the row number is in the same format as the rest of the module
			i = row - 1
		else:
			#Get the length of each row, and then choose the largest one
			lengths = list(map(len, allRows))
			i = np.argmax(lengths)
		item = allRows[i]

		#Determine if the length or the list of cell objects should be returned
		if (justLength):
			return len(item)

		#Determine if the cell contents or cell objects should be returned
		if (contents):
			#Loop through each cell and retrieve its contents
			cellContents = []
			for piece in item:
				contents = self.getCellValue(cell = piece)
				cellContents.append(contents)

			#Determine if None values should be replaced with "".
			if (noNone):
				for i, piece in enumerate(cellContents):
					if (piece == None):
						cellContents[i] = ""

			return cellContents
		return item

	def getCellWidth(self, row, column):
		"""Returns the [width, height] of a cell.

		row (int)     - The index of the row
		column (int)  - The index of the column. Can be a char

		Example Input: getCellWidth(1, 2)
		"""

		#Retrieve contents
		contents = self.getCellValue(row, column)

		#Record width of contents
		if (contents != None):
			width = len(str(contents))
			return width
		else:
			return None

	def setColumnWidth(self, column, newWidth = None):
		"""Changes the width of a column.

		column (int) - The index of the column. Can be a char
		newWidth (int) - The new width of the column. If None: auto adjust the width to the largest value in the column

		Example Input: setColumnWidth(3, 16)
		"""
		if (newWidth == None):
			#Find the longest cell in the column
			possibleWidths = []
			for i, row in enumerate(self.sheet.iter_rows()):
				width = self.getCellWidth(i + 1, column)

				#Record width of contents
				if (width != None):
					possibleWidths.append(width)

			#Compensate for blank column
			if (possibleWidths == []):
				newWidth = -1
			else:
				newWidth = max(possibleWidths)

		#Apply the new width
		newWidth += 2
		self.sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = newWidth

	def setColumnColor(self, column, color = "CCCCCC"):
		"""Changes the color of a column.

		column (int) - The index of the column. Can be a char
		color (str)    - The new color in hex format

		Example Input: setColumnColor("A")
		"""

		fillObject = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type="solid")

		for i, row in enumerate(self.sheet.iter_rows()):
			self.sheet[openpyxl.utils.get_column_letter(column) + str(i + 1)].fill = fillObject

	def saveBook(self, fileName, filePath = "", overlayOk = True, temporary = False, saveImages = True):
		"""Saves the workbook to a specified location.

		fileName (str)   - The name of the file
		filePath (str)   - Where the file is located
		overlayOk (bool) - If True: Images can overlap. If False: Any images under otehr ones will be deleted. If None: Images will be scooted to the right until they are ont under another
		temporary (bool) - If True: The file will be saved under the same name, but with "_temp" after it. For debugging things
		saveImages (bool) - If True: Images in the document will be preserved upon loading
			Images, charts, etc. are not read by openpyxl.
			In order to preserve images, charts, etc., each image is loaded and re-written into the loaded workbook
			Method for preservation from http://www.penwatch.net/cms/?p=582
			Help from: code.activestate.com/recipes/528870-class-for-writing-content-to-excel-and-formatting

		Example Input: saveBook("test")
		"""

		if (temporary):
			fileName += "_temp"

		try:
			#Ensure correct format
			if ("." not in fileName):
				fileName += ".xlsx"

			self.book.save(filePath + fileName)
		
		except IOError:
			#A book by that name is already open
			print("ERROR: The excel file is still open. The file has still been saved. Just close the current file without saving.")

	def openBook(self, fileName, filePath = "", readImages = False):
		"""User-friendly name for loadBook()."""

		self.loadBook(fileName, filePath = filePath, readImages = readImages)

	def loadBook(self, fileName, filePath = "", readImages = False):
		"""Loads a workbook from a specified location into memmory.

		fileName (str) - The name of the file
		filePath (str) - Where the file is located
		readImages (bool) - If True: Images in the document will be preserved upon loading
			Images, charts, etc. are not read by openpyxl.
			In order to preserve images, charts, etc., each image is loaded and re-written into the loaded workbook
			Method for preservation from http://www.penwatch.net/cms/?p=582
			Help from: code.activestate.com/recipes/528870-class-for-writing-content-to-excel-and-formatting

		Example Input: loadBook("test")
		"""

		#Ensure correct format
		if ("." not in fileName):
			fileName += ".xlsx"

		#Load the workbook into memory
		self.book = openpyxl.load_workbook(filePath + fileName)
		#print(filePath + fileName + ".xlsx", self.getAllSheetNames())
		self.changeSheet(0)

	def openFile(self, fileName, filePath ="./"):
		"""Opens the excel file for the user.

		fileName (str) - The name of the excel file

		openFile("converted")
		"""

		#Ensure correct format
		if ("." not in fileName):
			fileName += ".xlsx"

		try:
			os.startfile(fileName)
		except AttributeError:
			subprocess.call(['open', fileName])

	def test(self):
		#Load in the workbook
		self.loadBook("STANDARD_COMPONENTS_BOM-test", readImages = True)
		self.saveBook("test2")

if (__name__ == "__main__"):
	excel = Excel()
	excel.test()