import os
from auto_ms.modules import parse, process, gui
from auto_ms.modules import handle_input as hi
from pprint import pprint
from collections import defaultdict
import argparse
import pandas as pd

# Refactor this into process
import shutil
import openpyxl
from numpy import arange
from math import log


def tempsplit_entry():
    parser = argparse.ArgumentParser()
    parser.add_argument("data_file", type=str, help="MS data file to split by temperature")
    parser.add_argument("outdir", type=str, help="Output directory")
    args = parser.parse_args()
    tempsplit(args.data_file, args.outdir)


def tempsplit(target_path, outdir):
    outdir, target_path = os.path.normpath(outdir), os.path.normpath(target_path)
    parsed = parse.parse_file(target_path)
    header = parse.parse_header(parsed['header'])
    data = parse.parse_data(parsed['data'])
    split_data = process.split_data(data)

    # write the output
    if not os.path.exists(outdir):
        os.makedirs(outdir)
    if not os.path.exists(os.path.join(outdir, 'tsvs')):
        os.makedirs(os.path.join(outdir, 'tsvs'))
    for temp, df in split_data.items():
        print('Processing temperature: {}'.format(temp))
        df = process.subset(df)
        df.to_csv(os.path.join(outdir, 'tsvs', "T{}.tsv".format(temp)), 
                  sep="\t",
                  index=None,
                  quoting=3)
    print("Completed")
    return 0

def ika():
    '''
    Entry point
    '''
    parser = argparse.ArgumentParser()
    parser.add_argument("datafile", type=str, help="MS data file to split by temperature")
    parser.add_argument("out_path", type=str, help="Output path")
    parser.add_argument("-em", "--e_mass", type=str, help="Mass of E", default='')
    parser.add_argument("-ym", "--y_mass", type=str, help="Mass of Y complex", default='')
    parser.add_argument("-mo", "--moles", type=str, help="Moles of Sample", default='')
    parser.add_argument("-ma", "--mass", type=str, help="Mass of Sample", default='')
    parser.add_argument("-mm", "--mol_mass", type=str, help="Molecular mass of Sample", default='')
    parser.add_argument("-dmc", "--diamagnetic_correction", type=str, help="Diamagnetic Correction", default='')
    args = parser.parse_args()


    # Check inputs and generate/calculate as needed
    params = hi.ika_in(args)
    print('ika_in: ', params)
    print([params[i] for i in ["e_mass", "y_mass", "diamagnetic_correction", "moles"]])
    # pop a GUI if all input not provided
    if not all(params[i] != '' for i in ["e_mass", "y_mass", "diamagnetic_correction", "moles"]):
        fields = ["mass", "mol_mass", "e_mass", "moles", "diamagnetic_correction", "y_mass"]
        print('PARAMS: ', params)
        params = gui.param_input_loop(prefill=[params[i] for i in fields])
        # Handle that horrible keying of the GUI input field. REFACTOR THIS
        params = [params[i] for i in ['Mass_E', 'Mass_y', 'Diamagnetic_corr', 'moles']]

    # make all params floats
    for key, item in params.items():
        params[key] = float(item)

    e_mass, y_mass = params['e_mass'], params['y_mass']
    moles, dmc = params['moles'], params['diamagnetic_correction']
    
    # Parse the datafile and extract data
    parsed = parse.parse_file(args.datafile)
    data = parse.parse_data(parsed['data'])
    data = process.subset(data)
    data = process.split_data(data)

    # Calculate secondary values, optimize and get the total error squared and the x value
    processed = process.process_datafile(args.datafile,
                                         e_mass, y_mass, moles, dmc)
    print(processed)
    # Prepare output file
    shutil.copy('data/template.xlsx', args.out_path)
    wb = openpyxl.load_workbook(args.out_path)

    taus = []
    errors = []
    for temp in arange(5, 10.5, 0.5):

        error, (chi_t, chi_s, alpha, tau) = processed[temp]
        # Fill in raw data
        df = data[temp]
        sheet = wb.copy_worksheet(wb['Sheet1'])
        sheet.title = "ac_varT-{}K".format(temp)
        process.insert_df_excel(df, sheet, 5, 1, skip_n=1)
        # Insert fit params
        df = pd.DataFrame({"fit_params": [chi_t, chi_s, alpha, tau]})
        process.insert_df_excel(df, sheet, 5, 16, skip_n=1)
        # Fill in the baseline values (e_mass et al.)
        sheet.cell(row=3, column=2, value=moles)
        sheet.cell(row=3, column=4, value=dmc)
        sheet.cell(row=2, column=6, value=e_mass)
        sheet.cell(row=3, column=6, value=y_mass)
        # Record the tau, error value for next plot
        taus.append(tau)
        errors.append(error)

    # Create a new sheet for the tau values 
    sheet = wb.create_sheet('Arhenius_plot')
    df = pd.DataFrame({'err^2': errors,
                       'Tau': taus,
                       'T': [str(temp) for temp in arange(5, 10.5, 0.5)],
                       'Ln(tau)': [log(tau) for tau in taus],
                       '1/T': [1 / temp for temp in arange(5, 10.5, 0.5)]})
    df = df[['err^2', 'Tau', 'T', 'Ln(tau)', '1/T']]
    process.insert_df_excel(df, sheet, 2, 1)
    wb.save(args.out_path)






def execute(target_path, out_path):
    parsed = parse.parse_file(target_path)
    header = parse.parse_header(parsed['header'])
    data = parse.parse_data(parsed['data'])
    split_data = process.split_data(data)

    # write the output
    if not os.path.exists(out_path):
        os.makedirs(out_path)
    if not os.path.exists(os.path.join(out_path, 'tsvs')):
        os.makedirs(os.path.join(out_path, 'tsvs'))
    for temp, df in split_data.items():
        df = df[['Field (Oe)', 'Temperature (K)', "m' (emu)", 'm\" (emu)', 'Wave Frequency (Hz)']]
        df = process.calculate_chi(df, 51.2, 0, 1.80573e-05, -0.00049552)
        print(df.columns)
        df.to_csv(os.path.join(out_path, 'tsvs', "T{}.tsv".format(temp)), 
                  sep="\t",
                  index=None,
                  quoting=3)

if __name__ == '__main__':
    ika()

    # process.make_excel('testdata/d1.dat', 'output/test.xlsx')
    # execute('testdata/d1.dat', 'output/testout')
