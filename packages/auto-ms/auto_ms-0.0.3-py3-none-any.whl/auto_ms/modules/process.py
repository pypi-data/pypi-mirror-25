import pandas as pd
import shutil
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import math
import logging
from scipy.optimize import minimize
import numpy as np
from collections import defaultdict
from auto_ms.modules import parse


def split_data(data_df):
    '''
    splits the data based off a given key. Probably have to tweak this
    to make it more generic. handle the rounding outside?
    '''
    t_key = 'Temperature (K)'
    data_df = data_df
    data_df[t_key] = data_df[t_key].round(1)
    res = {}
    for temp in set(data_df[t_key].tolist()):
        res[temp] = data_df[data_df[t_key] == temp]

    return res

def subset(data_df):
    return data_df[['Field (Oe)', 'Temperature (K)', "m' (emu)", 'm\" (emu)', 'Wave Frequency (Hz)']]

def compute_chi(m, e_mass, y_mass,
                moles, dmc):
    """
    m', "" value
    eicosane mass
    Y complex mass
    moles
    diamagnetic correction
    """
    return ((m / 4) + (e_mass / 282548) * 0.00024306 +
            (y_mass / 698124) * 0.00040896) / moles - dmc


def calculate_chi(data_df, e_mass, y_mass,
                   moles, dmc):
    """
    Given a data_df containing m' and m" values, returns a dictionary
    containing two vectors with the calculated chi (emu) values
    """
    # calculate the emulated values. d1 is chi', d2 chi"
    d1 = compute_chi(data_df["m' (emu)"],
                     e_mass,
                     y_mass,
                     moles,
                     dmc)
    d2 = compute_chi(data_df['m" (emu)'],
                     e_mass,
                     y_mass,
                     moles,
                     dmc)
    return {"chi' (emu)": d1,
            'chi" (emu)': d2}

def compute_chi_p_fit(data_df,
                      chi_t, chi_s, alpha, tau):
    """
    Calculates the chi' (fit) value for a given wave_frequency value given a
    wave frequency vector
    """
    wave_freq = np.array(data_df['Wave Frequency (Hz)'])
    return chi_s + (chi_t - chi_s) * (1 + (2 * np.pi * wave_freq * tau) ** (1 - alpha) * np.sin(np.pi * alpha / 2)) / (1 + 2 * (2 * np.pi * wave_freq * tau) ** (1 - alpha) * np.sin(np.pi * alpha / 2) + (2 * np.pi * wave_freq * tau) ** (2 - 2 * alpha))

def compute_chi_pp_fit(data_df,
                      chi_t, chi_s, alpha, tau):
    wave_freq = np.array(data_df['Wave Frequency (Hz)'])
    return (chi_t - chi_s) * ((2 * np.pi * wave_freq * tau) ** (1 - alpha) * np.cos(np.pi * alpha / 2)) / (1 + 2 * (2 * np.pi * wave_freq * tau) ** (1 - alpha) * np.sin(np.pi * alpha / 2) + (2 * np.pi * wave_freq * tau) ** (2 - 2 * alpha))


def optimize_chi_fit(data_df, e_mass, y_mass, moles, dmc):
    '''
    takes values of form:
    (data_df, wave_freq, chi_t, chi_s, alpha, tau)
    '''
    chi_emu = calculate_chi(data_df, e_mass, y_mass, moles, dmc)
    chi_p, chi_pp = chi_emu["chi' (emu)"], chi_emu['chi" (emu)']

    # vectorize the chi_p emu and prepare data for optimization
    # Constraints
    # chi-t >= 0    (x[0])
    # chi-s >= 0    (x[1])
    # tau >= 0.0000001     (x[3])
    # cons = [
    #     {'type': 'ineq', 'fun': lambda x: x[0]},
    #     {'type': 'ineq', 'fun': lambda x: x[1]},
    #     {'type': 'ineq', 'fun': lambda x: x[2] - 0.0000001}
    # ]
    bnds = [(0, 10),
            (0, 10),
            (None, None),
            (0.0000001, None),]
    # take a vector of form ct, cs, a, t
    compute_p = lambda v : (sum((compute_chi_p_fit(data_df, *v) - chi_p) ** 2) +
                            sum((compute_chi_pp_fit(data_df, *v) - chi_pp) ** 2))
    res = minimize(compute_p, [0, 0, 0, 0], 
                   method='SLSQP', bounds=bnds,
                   options={'ftol':1E-50}) #constraints=cons)
    # logging.debug('Iterations: {}'.format(res.nit))
    return res


def process_datafile(target_path,
                     e_mass, y_mass, moles, dmc):
    '''
    Processes a datafile and calculates the chi emu and fits,
    spitting out all the desired values in a convenient excel file
    '''
    parsed = parse.parse_file(target_path)
    data = subset(parse.parse_data(parsed['data']))
    splat = split_data(data)
    chi = defaultdict(dict)
    res = {}
    for i in np.arange(5, 10.5, 0.5):
        chi_emu = calculate_chi(splat[i], e_mass, y_mass, moles, dmc)
        chi_p, chi_pp = chi_emu["chi' (emu)"], chi_emu['chi" (emu)']
        optimized = optimize_chi_fit(splat[i], e_mass, y_mass, moles, dmc)
        print(optimized.fun, optimized.x, i)
        res[i] = (optimized.fun, optimized.x)
    return res

def insert_df_excel(df, sheet, rt, ct, skip_n=0):
    for r, row in enumerate(dataframe_to_rows(df), start=rt):
        if skip_n and r < rt + skip_n:
            continue
        # We need to handle the 
        for c, col in enumerate(row, start= ct - 1):
            if c < ct:
                continue
            sheet.cell(row=r, column=c, value=col)
    return 1


def excel_from_template(out_path, 
                        e_mass, y_mass, moles, dmc,
                        chi_p = None, chi_pp = None,
                        optimized=None):
    """
    Given precomputed chi/optimized parameters, creates the output excel from a given template
    """
    wb = openpyxl.load_workbook(out_path)
    parsed = parse.parse_file(dat_path)

def make_excel(dat_path, out_path, optimized = None):
    shutil.copy('data/template.xlsx', out_path)
    wb = openpyxl.load_workbook(out_path)
    parsed = parse.parse_file(dat_path)
    header = parse.parse_header(parsed['header'])
    data = parse.parse_data(parsed['data'])
    data = process.subset(data)
    data = process.split_data(data)

    for temp in sorted(data):
        df = data[temp]
        # print(df)
        sheet = wb.copy_worksheet(wb['Sheet1'])
        sheet.title = "ac_varT-{}K".format(temp)
        insert_df_excel(df, sheet, 5, 1, skip_n=1)
        if optimized:
            insert_df_excel(df, sheet, 6, 16, skip_n=1)
    # wb.remove_sheet(wb['Sheet1'])
    wb.save(out_path)
